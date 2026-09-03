#!/usr/bin/env python3
"""
Verify that core/data.js's `const LT_910`/`const LT_911` lookup tables and
`const SUBNET_MASKS` list match the "Static Reference Tables" sheet of the
official Broadcom workbooks:
  - vcf-9.1-planning-and-preparation-workbook-25june2026.xlsx (v1.9.1.005) for LT_910
  - vcf-9.1.1-planning-and-preparation-workbook.xlsx (v1.9.1.101) for LT_911's
    overrides only (LT_911 = { ...LT_910, <overrides> } — unlisted keys are
    identical to LT_910 by construction and are already covered by that check)

(These tables used to live inline in index.html; they were extracted into the
shared ES module core/data.js — the single source of truth consumed by both
index.html and the MCP server in mcp/. The `export const LT_TABLES = ` /
`export const SUBNET_MASKS = ` declarations there still match the regexes below.)

Usage:
    python3 tools/check_lt_constants.py

Exits with code 0 if everything matches, 1 if any divergence is found
(and prints a ✓/✗ report), 2 if either workbook file cannot be found
(both are gitignored — keep copies next to index.html to run this check).

Requires: openpyxl (pip install openpyxl)
"""
import json
import os
import re
import sys
import warnings

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_FILE = os.path.join(ROOT, 'core', 'data.js')
EXCEL_FILE = os.path.join(ROOT, 'vcf-9.1-planning-and-preparation-workbook-25june2026.xlsx')  # v1.9.1.005, 25 Jun 2026
EXCEL_FILE_911 = os.path.join(ROOT, 'vcf-9.1.1-planning-and-preparation-workbook.xlsx')  # v1.9.1.101


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------

def extract_blocks(ws):
    """Walk the 'Lookup Tables' section (rows 31-350, columns B/C) and group
    consecutive tier/value rows under their preceding '<Name> | Value' header."""
    blocks = {}
    current = None
    for row in range(31, 351):
        b = ws.cell(row=row, column=2).value
        c = ws.cell(row=row, column=3).value
        if c == 'Value' and b:
            current = str(b).strip()
            blocks[current] = {}
            continue
        if current is not None and b is not None and isinstance(c, (int, float)):
            blocks[current][str(b).strip()] = c
    return blocks


def extract_subnet_masks(ws):
    masks = []
    for row in range(320, 353):
        c = ws.cell(row=row, column=3).value
        if c is not None:
            masks.append(str(c))
    return masks


def tri(blocks, cpu_block, ram_block, disk_block, tier_map=None):
    cpu, ram, disk = blocks[cpu_block], blocks[ram_block], blocks[disk_block]
    tiers = set(cpu) & set(ram) & set(disk)
    out = {}
    for t in tiers:
        lt_tier = (tier_map or {}).get(t, t)
        out[lt_tier] = {'vcpu': cpu[t], 'ram': ram[t], 'disk': disk[t]}
    return out


def parse_vcenter_disk(blocks):
    """Parse the single 'vCenter Disk' block, whose keys concatenate a vCenter
    appliance size (Tiny/Small/Medium/Large/XLarge) with a disk tier
    (Default/Large/XLarge, plus 9.1.1's lstorage/xlstorage aliases) with no
    separator, e.g. 'TinyDefault', 'XLargeLarge', 'Tinyxlstorage'. Note 'Large' is
    ambiguous (appears as both a size and a tier name, and 'xlstorage' ends with
    'lstorage'), so suffixes are tested longest/most-specific first (XLarge before
    Large, xlstorage before lstorage) and stripped from the key to recover the
    size."""
    raw = blocks['vCenter Disk']
    tiers = ('Default', 'XLarge', 'xlstorage', 'Large', 'lstorage')
    out = {}
    for key, value in raw.items():
        for tier in tiers:
            if key.endswith(tier):
                size = key[: -len(tier)]
                break
        else:
            raise ValueError(f"vCenter Disk: unrecognized tier suffix in key {key!r}")
        # The 9.1.1 workbook has one inconsistently-cased row ('Xlargexlstorage'
        # instead of 'XLargexlstorage') — fold case-insensitively against the known
        # size names rather than treating it as a distinct 6th size.
        for known_size in ('Tiny', 'Small', 'Medium', 'Large', 'XLarge'):
            if size.lower() == known_size.lower():
                size = known_size
                break
        out.setdefault(size, {})[tier] = value
    return out


def build_expected(blocks):
    e = {}
    e['sddc_manager'] = {
        'vcpu': blocks['SDDC Manager']['CPU'],
        'ram': blocks['SDDC Manager']['RAM'],
        'disk': blocks['SDDC Manager']['Disk'],
    }
    e['nsx_manager'] = tri(blocks, 'NSX-T Manager CPU', 'NSX-T Manager RAM', 'NSX-T Manager Disk')
    e['nsx_edge'] = tri(blocks, 'NSX-T Edge CPU', 'NSX-T Edge RAM', 'NSX-T Edge Disk')
    e['nsx_edge']['Excluded'] = {'vcpu': 0, 'ram': 0, 'disk': 0}
    e['vcf_operations'] = tri(blocks, 'VCF Operations CPU', 'VCF Operations RAM', 'VCF Operations Disk',
                               tier_map={'Extra Small': 'Extra-Small', 'Extra Large': 'Extra-Large'})
    e['avi_lb'] = tri(blocks, 'AVI Load Balancer CPU', 'AVI Load Balancer Ram', 'AVI Load Balancer Disk',
                      tier_map={'X-Large': 'Extra-Large'})
    e['vcfops_proxy'] = tri(blocks, 'VCF Operations Proxy CPU', 'VCF Operations Proxy RAM', 'VCF Operations Proxy Disk')
    e['ssp'] = tri(blocks, 'SSP CPU', 'SSP RAM', 'SSP Disk')
    e['ssp']['Excluded'] = {'vcpu': 0, 'ram': 0, 'disk': 0}
    e['ssp_license'] = {
        'vcpu': blocks['SSP License']['CPU'],
        'ram': blocks['SSP License']['RAM'],
        'disk': blocks['SSP License']['Storage'],
    }
    e['vcfms_control_node'] = tri(blocks, 'VCFMS Control Node CPU', 'VCFMS Control Node RAM', 'VCFMS Control Node Disk')
    e['vcfms_worker_node'] = tri(blocks, 'VCFMS Worker Node CPU', 'VCFMS Worker Node RAM', 'VCFMS Worker Node Disk')
    e['vcfms_control_nodes'] = dict(blocks['Deployment Size Control Node'])
    e['vcfms_worker_nodes'] = dict(blocks['VCFMS Worker Node'])
    e['identity_broker'] = tri(blocks, 'Identity Broker CPU', 'Identity Broker RAM', 'Identity Broker Disk')
    e['vcf_ops_networks'] = tri(blocks, 'VCF Operations for networks CPU', 'VCF Operations for networks RAM',
                                 'VCF Operations for networks DISK')
    e['vcf_ops_networks_collector'] = tri(
        blocks,
        'VCF Operations for networks - Collector CPU',
        'VCF Operations for networks - Collector RAM',
        'VCF Operations for networks - Collecter DISK',
        tier_map={'Extra Large': 'Extra-Large', 'Extra Extra Large': 'Extra-Extra-Large'},
    )
    e['vrms'] = tri(blocks, 'VRMS CPU', 'VRMS RAM', 'VRMS Disk')
    e['srm'] = tri(blocks, 'SRM CPU', 'SRM RAM', 'SRM Disk')
    e['hvm'] = {
        'vcpu': blocks['Health Reporting and Monitoring - HVM']['CPU'],
        'ram': blocks['Health Reporting and Monitoring - HVM']['RAM'],
        'disk': blocks['Health Reporting and Monitoring - HVM']['Disk'],
    }
    e['cloud_ransomware'] = {
        'vcpu': blocks['Cloud-Based Ransomware']['CPU'],
        'ram': blocks['Cloud-Based Ransomware']['RAM'],
        'disk': blocks['Cloud-Based Ransomware']['Disk'],
    }
    e['hcx_connector'] = {
        'vcpu': blocks['Cross-Cloud Mobility - HCX Conn']['CPU'],
        'ram': blocks['Cross-Cloud Mobility - HCX Conn']['RAM'],
        'disk': blocks['Cross-Cloud Mobility - HCX Conn']['Disk'],
    }
    e['vcenter_disk_tiers'] = parse_vcenter_disk(blocks)
    e['vcfa'] = tri(blocks, 'VCF Automation CPU', 'VCF Automation RAM', 'VCF Automation Disk')
    return e


# ---------------------------------------------------------------------------
# core/data.js extraction (minimal JS-object-literal -> JSON converter)
# ---------------------------------------------------------------------------

def _extract_balanced(text, start):
    """Given text and the index of an opening '{' or '[', return the matching
    substring up to (and including) the balanced closing bracket.

    Uses naive bracket counting (ignores quoting/comments) — safe here
    because every brace/bracket that appears inside a string or comment in
    the LT/SUBNET_MASKS blocks is part of a balanced pair (e.g. the comment
    "{vcpu,ram,disk}"), so it nets to zero and doesn't shift the depth count.
    """
    open_ch = text[start]
    close_ch = {'{': '}', '[': ']'}[open_ch]
    depth = 0
    i = start
    while i < len(text):
        ch = text[i]
        if ch == open_ch:
            depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
        i += 1
    raise ValueError('unbalanced brackets')


def _js_literal_to_json(src):
    # Strip // line comments (safe here: no string values contain "//")
    src = re.sub(r'//[^\n]*', '', src)
    # Single-quoted strings -> double-quoted
    src = re.sub(r"'((?:[^'\\]|\\.)*)'", lambda m: json.dumps(m.group(1)), src)
    # Quote bare identifier keys: { foo: ... or , foo: ...
    src = re.sub(r'([{,]\s*)([A-Za-z_][A-Za-z0-9_]*)\s*:', r'\1"\2":', src)
    # Remove trailing commas before } or ]
    src = re.sub(r',\s*([}\]])', r'\1', src)
    return src


def extract_lt_and_masks(html_text):
    """Returns (lt_910, lt_911_overrides, masks). lt_911_overrides holds ONLY the
    keys core/data.js's `const LT_911 = { ...LT_910, <overrides> }` actually
    overrides — the `...LT_910` spread isn't valid JSON, so it's stripped rather
    than resolved; unlisted keys are identical to LT_910 and are covered by that
    comparison already."""
    m = re.search(r'const LT_910 = ', html_text)
    lt_910_raw = _extract_balanced(html_text, m.end())
    lt_910 = json.loads(_js_literal_to_json(lt_910_raw))

    m = re.search(r'const LT_911 = ', html_text)
    lt_911_raw = _extract_balanced(html_text, m.end())
    lt_911_raw = lt_911_raw.replace('...LT_910,', '', 1)
    lt_911_overrides = json.loads(_js_literal_to_json(lt_911_raw))

    m = re.search(r'const SUBNET_MASKS = ', html_text)
    masks_raw = _extract_balanced(html_text, m.end())
    masks = json.loads(_js_literal_to_json(masks_raw))

    return lt_910, lt_911_overrides, masks


# ---------------------------------------------------------------------------
# Comparison / report
# ---------------------------------------------------------------------------

def compare(expected, actual, label, ok, problems):
    if actual is None:
        problems.append(f"{label}: MISSING from core/data.js LT")
        print(f"  ✗ {label}: missing")
        return False
    all_ok = True
    if isinstance(expected, dict) and all(isinstance(v, dict) for v in expected.values()):
        # tiered component
        for tier, vals in expected.items():
            a = actual.get(tier)
            if a is None:
                problems.append(f"{label}[{tier}]: missing tier in core/data.js")
                print(f"  ✗ {label}[{tier}]: missing tier")
                all_ok = False
                continue
            for k, v in vals.items():
                if a.get(k) != v:
                    problems.append(f"{label}[{tier}].{k}: expected {v}, got {a.get(k)}")
                    print(f"  ✗ {label}[{tier}].{k}: expected {v}, got {a.get(k)}")
                    all_ok = False
    else:
        for k, v in expected.items():
            if actual.get(k) != v:
                problems.append(f"{label}.{k}: expected {v}, got {actual.get(k)}")
                print(f"  ✗ {label}.{k}: expected {v}, got {actual.get(k)}")
                all_ok = False
    if all_ok:
        print(f"  ✓ {label}")
    return all_ok


def compare_vcenter_disk_tiers(expected, actual, problems):
    label = 'vcenter_disk_tiers'
    if actual is None:
        problems.append(f"{label}: MISSING from core/data.js LT")
        print(f"  ✗ {label}: missing")
        return False
    all_ok = True
    for size, tiers in expected.items():
        a = actual.get(size)
        if a is None:
            problems.append(f"{label}[{size}]: missing size in core/data.js")
            print(f"  ✗ {label}[{size}]: missing size")
            all_ok = False
            continue
        for tier, v in tiers.items():
            if a.get(tier) != v:
                problems.append(f"{label}[{size}].{tier}: expected {v}, got {a.get(tier)}")
                print(f"  ✗ {label}[{size}].{tier}: expected {v}, got {a.get(tier)}")
                all_ok = False
    if all_ok:
        print(f"  ✓ {label}")
    return all_ok


# 9.1.1-changed keys only — everything else in LT_911 is identical to LT_910 (a JS
# spread, not re-typed), so only these need re-checking against the 9.1.1 workbook.
LT_911_CHANGED_KEYS = {'ssp', 'ssp_license', 'vcfops_proxy', 'vcf_ops_networks', 'vcf_ops_networks_collector', 'vcenter_disk_tiers'}


def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"Workbook not found: {EXCEL_FILE}")
        print("(it's gitignored — place a copy next to index.html to run this check)")
        return 2
    if not os.path.exists(EXCEL_FILE_911):
        print(f"Workbook not found: {EXCEL_FILE_911}")
        print("(it's gitignored — place a copy next to index.html to run this check)")
        return 2

    import openpyxl
    warnings.filterwarnings('ignore')

    with open(SOURCE_FILE, encoding='utf-8') as f:
        source_text = f.read()
    lt_910, lt_911_overrides, masks = extract_lt_and_masks(source_text)

    problems = []

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Static Reference Tables']
    blocks = extract_blocks(ws)
    expected_910 = build_expected(blocks)
    expected_masks = extract_subnet_masks(ws)

    print(f'Comparing core/data.js `LT_910` against 9.1.0 workbook ({os.path.basename(EXCEL_FILE)})...')
    for key in expected_910:
        if key == 'vcenter_disk_tiers':
            continue
        compare(expected_910[key], lt_910.get(key), key, True, problems)
    compare_vcenter_disk_tiers(expected_910['vcenter_disk_tiers'], lt_910.get('vcenter_disk_tiers'), problems)

    wb911 = openpyxl.load_workbook(EXCEL_FILE_911, data_only=True)
    ws911 = wb911['Static Reference Tables']
    blocks911 = extract_blocks(ws911)
    # NOT build_expected(blocks911) — that also computes vcfms_worker_node/etc.,
    # whose 9.1.1 source blocks moved past extract_blocks()'s fixed row-31-350
    # window (9.1.1's VCFMS worker-node model was replaced entirely anyway, see
    # core/data.js's VCFMS_911 — this script intentionally doesn't check it).
    expected_911 = {
        'ssp': tri(blocks911, 'SSP CPU', 'SSP RAM', 'SSP Disk'),
        'ssp_license': {
            'vcpu': blocks911['SSP License']['CPU'],
            'ram': blocks911['SSP License']['RAM'],
            'disk': blocks911['SSP License']['Storage'],
        },
        'vcfops_proxy': tri(blocks911, 'VCF Operations Proxy CPU', 'VCF Operations Proxy RAM', 'VCF Operations Proxy Disk'),
        'vcf_ops_networks': tri(blocks911, 'VCF Operations for networks CPU', 'VCF Operations for networks RAM',
                                 'VCF Operations for networks DISK'),
        'vcf_ops_networks_collector': tri(
            blocks911,
            'VCF Operations for networks - Collector CPU',
            'VCF Operations for networks - Collector RAM',
            'VCF Operations for networks - Collecter DISK',
        ),
        'vcenter_disk_tiers': parse_vcenter_disk(blocks911),
    }
    expected_911['ssp']['Excluded'] = {'vcpu': 0, 'ram': 0, 'disk': 0}

    print(f'\nComparing core/data.js `LT_911` overrides against 9.1.1 workbook ({os.path.basename(EXCEL_FILE_911)})...')
    if set(lt_911_overrides) != LT_911_CHANGED_KEYS:
        problems.append(f'LT_911 override keys changed: expected {sorted(LT_911_CHANGED_KEYS)}, got {sorted(lt_911_overrides)}')
        print(f"  ✗ LT_911 now overrides {sorted(lt_911_overrides)} — update LT_911_CHANGED_KEYS and this script's expectations")
    for key in LT_911_CHANGED_KEYS:
        if key == 'vcenter_disk_tiers':
            continue
        compare(expected_911[key], lt_911_overrides.get(key), key, True, problems)
    compare_vcenter_disk_tiers(expected_911['vcenter_disk_tiers'], lt_911_overrides.get('vcenter_disk_tiers'), problems)

    print('\nComparing SUBNET_MASKS (9.1.0 workbook)...')
    if masks == expected_masks:
        print(f"  ✓ SUBNET_MASKS ({len(masks)} entries)")
    else:
        problems.append('SUBNET_MASKS mismatch')
        print(f"  ✗ SUBNET_MASKS: expected {len(expected_masks)} entries, got {len(masks)}")
        for i, (e, a) in enumerate(zip(expected_masks, masks)):
            if e != a:
                print(f"      index {i}: expected {e!r}, got {a!r}")
        if len(masks) != len(expected_masks):
            print(f"      expected: {expected_masks}")
            print(f"      actual:   {masks}")

    print()
    if problems:
        print(f"FAILED — {len(problems)} divergence(s) found.")
        return 1
    print("OK — core/data.js LT_910/LT_911/SUBNET_MASKS match both workbooks.")
    return 0


if __name__ == '__main__':
    sys.exit(main())
